from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")

MST_PATTERN = re.compile(r"\d{10}(?:-\d{3})?")
DEFAULT_GDT_SOURCE = "https://tracuunnt.gdt.gov.vn/tcnnt/mstdn.jsp"
OUTPUT_COLUMNS = ["STT", "MST", "Ten nguoi nop thue", "Dia chi", "Co quan thue quan ly", "Trang thai MST"]
CAPTCHA_ERROR_MARKERS = ("nhap dung ma xac nhan", "ma xac nhan khong dung", "captcha khong dung")
NOT_FOUND_MARKERS = ("khong tim thay", "khong co thong tin")


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFD", value or "")
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    text = re.sub(r"[^0-9A-Za-z]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def coerce_mst(value: object, numeric_from_excel: bool = False) -> Optional[str]:
    if value is None:
        return None
    if numeric_from_excel and isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return None
        text = str(int(value)).zfill(10)
    else:
        text = str(value).strip().replace("\ufeff", "")
    match = MST_PATTERN.search(re.sub(r"[^\d-]", "", text))
    return match.group(0) if match else None


def normalize_mst(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def unique_keep_order(values: Iterable[str]) -> List[str]:
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def read_text_or_csv(path: Path) -> List[str]:
    msts: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        if path.suffix.lower() == ".csv":
            reader = csv.reader(handle)
            for row in reader:
                for cell in row:
                    mst = coerce_mst(cell)
                    if mst:
                        msts.append(mst)
        else:
            for line in handle:
                mst = coerce_mst(line)
                if mst:
                    msts.append(mst)
    return msts


def resolve_excel_column(sheet, column: Optional[str]) -> Optional[int]:
    if not column:
        return None
    if column.isdigit():
        index = int(column)
        if index <= 0:
            raise ValueError("--input-column must be a 1-based column index or header name")
        return index
    target = normalize_text(column)
    for cell in next(sheet.iter_rows(min_row=1, max_row=1), []):
        if normalize_text(str(cell.value or "")) == target:
            return cell.column
    raise ValueError(f"Could not find Excel column header: {column}")


def read_excel(path: Path, input_column: Optional[str]) -> List[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    column_index = resolve_excel_column(sheet, input_column)
    msts: List[str] = []
    for row in sheet.iter_rows():
        cells = row if column_index is None else [row[column_index - 1]]
        for cell in cells:
            mst = coerce_mst(cell.value, numeric_from_excel=True)
            if mst:
                msts.append(mst)
    workbook.close()
    return msts


def read_msts_from_file(path: Path, input_column: Optional[str]) -> List[str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel(path, input_column)
    if suffix in {".txt", ".csv"}:
        return read_text_or_csv(path)
    raise ValueError("Input file must be .txt, .csv, .xlsx, or .xlsm")


def get_msts(args: argparse.Namespace) -> List[str]:
    if args.input_file:
        values = read_msts_from_file(Path(args.input_file), args.input_column)
    elif len(args.inputs) == 1 and Path(args.inputs[0]).exists():
        values = read_msts_from_file(Path(args.inputs[0]), args.input_column)
    elif args.inputs:
        values = [mst for item in args.inputs if (mst := coerce_mst(item))]
    elif not sys.stdin.isatty():
        values = [mst for line in sys.stdin for mst in [coerce_mst(line)] if mst]
    else:
        raise ValueError("Provide MST values or an input file")
    if not args.keep_duplicates:
        values = unique_keep_order(values)
    if not values:
        raise ValueError("No valid MST values found")
    return values


def load_gdt_client() -> Tuple[Any, str]:
    try:
        from api_client_v2 import GdtTaxLookupClientV2, TAX_LOOKUP_URL
    except ImportError as exc:
        raise RuntimeError("Could not import api_client_v2") from exc
    return GdtTaxLookupClientV2, TAX_LOOKUP_URL or DEFAULT_GDT_SOURCE


def set_log_level(debug: bool) -> None:
    level = logging.INFO if debug else logging.ERROR
    logging.getLogger().setLevel(level)
    logging.getLogger("api_client_v2").setLevel(level)


def html_has_marker(html: Optional[str], markers: Iterable[str]) -> bool:
    normalized = normalize_text(html or "")
    return any(marker in normalized for marker in markers)


def select_taxpayer(mst: str, taxpayers: List[Any]) -> Optional[Any]:
    if not taxpayers:
        return None
    target = normalize_mst(mst)
    for taxpayer in taxpayers:
        if normalize_mst(getattr(taxpayer, "tax_id", "")) == target:
            return taxpayer
    return taxpayers[0]


def taxpayer_to_result(input_mst: str, taxpayer: Any, source_url: str) -> Dict[str, str]:
    return {
        "input_mst": input_mst,
        "mst": getattr(taxpayer, "tax_id", input_mst),
        "company_name": getattr(taxpayer, "name", ""),
        "address": getattr(taxpayer, "address", ""),
        "managed_by": getattr(taxpayer, "tax_office", ""),
        "status": getattr(taxpayer, "status", ""),
        "source_url": source_url,
    }


def save_html_path(save_html_dir: Optional[str], mst: str, attempt: int) -> Optional[str]:
    if not save_html_dir:
        return None
    html_dir = Path(save_html_dir)
    html_dir.mkdir(parents=True, exist_ok=True)
    safe_mst = re.sub(r"[^0-9-]", "_", mst)
    return str(html_dir / f"{safe_mst}_attempt_{attempt}.html")


def lookup_mst_with_client(client: Any, mst: str, retries: int, source_url: str, save_html_dir: Optional[str]) -> Dict[str, str]:
    try:
        if not client.init_session():
            detail = getattr(client, "last_error", "") or ""
            message = "Could not initialize GDT session"
            if detail:
                message = f"{message}: {detail}"
            return {"input_mst": mst, "mst": mst, "error": message, "source_url": source_url}
    except Exception as exc:
        return {"input_mst": mst, "mst": mst, "error": f"Could not initialize GDT session: {exc}", "source_url": source_url}

    for attempt in range(1, retries + 1):
        try:
            captcha = client.auto_solve_captcha()
            print(f"  CAPTCHA attempt {attempt}/{retries}: {captcha}")
            taxpayers = client.lookup_tax_id(mst, captcha, save_html_path=save_html_path(save_html_dir, mst, attempt))
            selected = select_taxpayer(mst, taxpayers)
            if selected:
                return taxpayer_to_result(mst, selected, source_url)
            html = getattr(client, "last_html", "")
            if html_has_marker(html, NOT_FOUND_MARKERS):
                return {"input_mst": mst, "mst": mst, "error": "Not found", "source_url": source_url}
            if attempt < retries:
                reason = "wrong CAPTCHA" if html_has_marker(html, CAPTCHA_ERROR_MARKERS) else "no result yet"
                print(f"  -> {reason}, retrying...")
                time.sleep(1)
                continue
            return {"input_mst": mst, "mst": mst, "error": "No result after CAPTCHA retries", "source_url": source_url}
        except Exception as exc:
            if attempt == retries:
                return {"input_mst": mst, "mst": mst, "error": str(exc), "source_url": source_url}
            print(f"  -> attempt {attempt}/{retries} error: {exc}. Retrying...")
            time.sleep(1)
            try:
                client.init_session()
            except Exception:
                pass
    return {"input_mst": mst, "mst": mst, "error": "Lookup failed", "source_url": source_url}


def build_row(index: int, result: Dict[str, str]) -> List[object]:
    status = result.get("status") or result.get("error") or ""
    return [index, result.get("mst") or result.get("input_mst", ""), result.get("company_name", ""), result.get("address", ""), result.get("managed_by", ""), status]


def write_excel(results: List[Dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tra cuu thue"
    sheet.append(OUTPUT_COLUMNS)
    for index, result in enumerate(results, start=1):
        sheet.append(build_row(index, result))
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment
    for column, width in {"A": 8, "B": 18, "C": 42, "D": 58, "E": 36, "F": 28}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch lookup Vietnamese tax IDs and export Excel results")
    parser.add_argument("inputs", nargs="*", help="MST values or one .txt/.csv/.xlsx/.xlsm input file")
    parser.add_argument("-i", "--input-file", help="Input file containing MST values")
    parser.add_argument("--input-column", help="Excel column index or header name containing MST values")
    parser.add_argument("-o", "--output", default="tra_cuu_thue_result.xlsx", help="Output Excel file")
    parser.add_argument("--delay", type=float, default=1.0, help="Delay between MST lookups in seconds")
    parser.add_argument("--retries", type=int, default=15, help="CAPTCHA retry count per MST")
    parser.add_argument("--save-html-dir", help="Directory for debug HTML responses")
    parser.add_argument("--debug", action="store_true", help="Enable detailed logging")
    parser.add_argument("--keep-duplicates", action="store_true", help="Keep duplicate MST values")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    set_log_level(args.debug)
    try:
        msts = get_msts(args)
        client_class, source_url = load_gdt_client()
    except Exception as exc:
        print(f"Input error: {exc}", file=sys.stderr)
        return 2

    results: List[Dict[str, str]] = []
    total = len(msts)
    with client_class() as client:
        for index, mst in enumerate(msts, start=1):
            print(f"[{index}/{total}] Looking up MST {mst}...")
            result = lookup_mst_with_client(client, mst, args.retries, source_url, args.save_html_dir)
            results.append(result)
            if result.get("error"):
                print(f"  -> ERROR: {result['error']}")
            else:
                print(f"  -> OK: {result.get('company_name', '')}")
            if index < total and args.delay > 0:
                time.sleep(args.delay)

    output_path = Path(args.output)
    write_excel(results, output_path)
    print(f"Wrote Excel: {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
